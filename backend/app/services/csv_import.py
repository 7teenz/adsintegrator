from __future__ import annotations

import csv
import hashlib
import io
import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from typing import Any

from fastapi import HTTPException, status
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from sqlalchemy.orm import Session

from app.models.campaign import Ad, AdSet, Campaign, Creative
from app.models.insights import DailyAccountInsight, DailyAdInsight, DailyAdSetInsight, DailyCampaignInsight
from app.models.meta_connection import MetaAdAccount, MetaConnection
from app.models.user import User
from app.services.crypto import encrypt_token
from app.services.meta_ads import MetaAdsService


@dataclass
class CsvImportResult:
    campaigns: int
    ad_sets: int
    ads: int
    insight_rows: int
    date_start: date | None
    date_end: date | None
    report_type: str = "daily_breakdown"
    source_sheet: str | None = None
    warnings: list[str] = field(default_factory=list)


class CsvImportService:
    SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}
    ALLOWED_MIME_TYPES = {
        "text/plain",
        "text/csv",
        "application/csv",
        "application/vnd.ms-excel",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/octet-stream",
        "application/zip",
    }

    COLUMN_ALIASES = {
        "date": ["date", "day", "date_start", "reporting starts", "дата", "дата начала отчетности", "начало"],
        "date_end": ["date_end", "reporting ends", "дата окончания отчетности", "конец"],
        "campaign_id": ["campaign id", "campaign_id", "идентификатор кампании"],
        "campaign_name": ["campaign name", "campaign_name", "campaign", "название кампании"],
        "adset_id": ["ad set id", "adset id", "ad_set_id", "adset_id", "идентификатор группы объявлений"],
        "adset_name": ["ad set name", "adset name", "ad_set_name", "adset_name", "название группы объявлений"],
        "ad_id": ["ad id", "ad_id", "идентификатор объявления"],
        "ad_name": ["ad name", "ad_name", "название объявления"],
        "spend": ["amount spent", "amount spent (usd)", "amount spent (eur)", "spend", "сумма затрат", "сумма затрат (usd)", "расход"],
        "impressions": ["impressions", "показы"],
        "reach": ["reach", "охват"],
        "clicks": ["clicks", "link clicks", "clicks (all)", "клики", "клики по ссылке"],
        "ctr": ["ctr", "ctr (all)", "ctr (link click-through rate)", "ctr (все)", "ctr (кликабельность всех элементов)"],
        "cpc": ["cpc", "cpc (all)", "cost per click", "cpc (цена за клик по ссылке)"],
        "cpm": ["cpm", "cpm (cost per 1,000 impressions)", "cost per 1,000 impressions", "cpm (цена за 1 000 показов)"],
        "frequency": ["frequency", "частота"],
        "conversions": ["results", "result", "conversions", "purchases", "website purchases", "результат", "результаты", "конверсии"],
        "conversion_value": ["purchase conversion value", "website purchase conversion value", "conversion value", "value", "ценность конверсий"],
        "roas": ["purchase roas", "website purchase roas", "roas", "окупаемость затрат на рекламу"],
        "result_type": ["result type", "тип результата"],
        "status": ["delivery", "status", "статус показа"],
        "objective": ["objective", "цель"],
        "level": ["level", "delivery level", "уровень показа"],
    }

    @classmethod
    def import_report(
        cls,
        db: Session,
        user: User,
        filename: str,
        content_bytes: bytes,
        replace_existing: bool = False,
    ) -> CsvImportResult:
        extension = cls._detect_extension(filename)
        if extension not in cls.SUPPORTED_EXTENSIONS:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"detail": "Only CSV and XLSX uploads are supported", "code": "REPORT_INVALID_FILE_TYPE"},
            )
        if len(content_bytes) > 10 * 1024 * 1024:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"detail": "Report file is too large (max 10MB)", "code": "REPORT_FILE_TOO_LARGE"},
            )
        detected_mime = cls._detect_mime_from_bytes(content_bytes)
        if detected_mime not in cls.ALLOWED_MIME_TYPES:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"detail": f"File type not allowed: {detected_mime}", "code": "REPORT_INVALID_MIME"},
            )

        account = cls._ensure_selected_account(db, user)
        rows, metadata = cls._read_rows(content_bytes, extension)
        if not rows:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"detail": "Report file has no data rows", "code": "REPORT_EMPTY"},
            )

        normalized_rows = [cls._normalize_row(row) for row in rows]
        normalized_rows = [row for row in normalized_rows if row["date"] is not None]
        normalized_rows = [
            row
            for row in normalized_rows
            if row["campaign_name"] or row["adset_name"] or row["ad_name"] or row["spend"] > 0 or row["impressions"] > 0
        ]
        if not normalized_rows:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={
                    "detail": "Could not detect a supported Meta report layout. Make sure the file contains recognizable headers and reporting dates.",
                    "code": "REPORT_INVALID_SCHEMA",
                },
            )

        if replace_existing:
            cls._clear_account_data(db, account.id)

        campaign_map = cls._upsert_campaigns(db, account.id, normalized_rows)
        adset_map = cls._upsert_adsets(db, account.id, normalized_rows, campaign_map)
        cls._upsert_ads(db, account.id, normalized_rows, adset_map)
        insight_rows = cls._upsert_insights(db, account.id, normalized_rows, campaign_map, adset_map)
        db.commit()

        date_values = [row["date"] for row in normalized_rows if row["date"] is not None]
        return CsvImportResult(
            campaigns=len(campaign_map),
            ad_sets=len(adset_map),
            ads=db.query(Ad).filter(Ad.ad_account_id == account.id).count(),
            insight_rows=insight_rows,
            date_start=min(date_values) if date_values else None,
            date_end=max(date_values) if date_values else None,
            report_type=metadata.get("report_type", "daily_breakdown"),
            source_sheet=metadata.get("source_sheet"),
            warnings=metadata.get("warnings", []),
        )

    @classmethod
    def import_csv(
        cls,
        db: Session,
        user: User,
        filename: str,
        content_bytes: bytes,
        replace_existing: bool = False,
    ) -> CsvImportResult:
        return cls.import_report(db=db, user=user, filename=filename, content_bytes=content_bytes, replace_existing=replace_existing)

    @staticmethod
    def _ensure_selected_account(db: Session, user: User) -> MetaAdAccount:
        connection = db.query(MetaConnection).filter(MetaConnection.user_id == user.id).first()
        if connection is None:
            connection = MetaConnection(
                user_id=user.id,
                meta_user_id="report_import",
                meta_user_name=user.full_name or user.email,
                encrypted_access_token=encrypt_token("report-import-local"),
                scopes="report_import",
            )
            db.add(connection)
            db.flush()
            account = MetaAdAccount(
                connection_id=connection.id,
                account_id=f"import_{user.id[:8]}",
                account_name="Imported Report Account",
                currency="USD",
                timezone="UTC",
                is_selected=True,
            )
            db.add(account)
            db.commit()
            db.refresh(account)
            return account

        account = MetaAdsService.get_selected_account(db, connection.id)
        if account is not None:
            return account

        accounts = MetaAdsService.get_ad_accounts(db, connection.id)
        if accounts:
            return MetaAdsService.select_ad_account(db, connection.id, accounts[0].account_id) or accounts[0]

        created = MetaAdAccount(
            connection_id=connection.id,
            account_id=f"import_{user.id[:8]}",
            account_name="Imported Report Account",
            currency="USD",
            timezone="UTC",
            is_selected=True,
        )
        db.add(created)
        db.commit()
        db.refresh(created)
        return created

    @classmethod
    def _read_rows(cls, content_bytes: bytes, extension: str) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        if extension == ".xlsx":
            return cls._read_xlsx_rows(content_bytes)
        return cls._read_csv_rows(content_bytes)

    @classmethod
    def _read_csv_rows(cls, content_bytes: bytes) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        text = None
        for encoding in ("utf-8-sig", "utf-8", "cp1251", "cp1252", "latin-1"):
            try:
                text = content_bytes.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"detail": "Could not decode report file", "code": "REPORT_DECODE_FAILED"},
            )

        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.get_dialect("excel")

        reader = csv.DictReader(io.StringIO(text), dialect=dialect)
        rows = [row for row in reader if row]
        return rows, {"report_type": cls._classify_report_type(rows), "source_sheet": None, "warnings": []}

    @classmethod
    def _read_xlsx_rows(cls, content_bytes: bytes) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        try:
            workbook = load_workbook(io.BytesIO(content_bytes), data_only=True, read_only=True)
        except Exception as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"detail": "Could not read XLSX file", "code": "REPORT_DECODE_FAILED"},
            ) from exc

        best_records: list[dict[str, Any]] = []
        best_sheet: str | None = None
        best_score = -1
        for candidate in workbook.worksheets:
            records, score = cls._extract_records_from_sheet(candidate)
            title = candidate.title.lower()
            if "raw" in title:
                score += 3
            if "formatted" in title:
                score -= 1
            if records and score > best_score:
                best_records = records
                best_sheet = candidate.title
                best_score = score

        if not best_records:
            return [], {"report_type": "unknown", "source_sheet": None, "warnings": ["No recognizable worksheet found."]}

        report_type = cls._classify_report_type(best_records)
        warnings = ["Aggregate report detected. Trend and anomaly sections will be limited."] if report_type == "period_aggregate" else []
        return best_records, {"report_type": report_type, "source_sheet": best_sheet, "warnings": warnings}

    @classmethod
    def _extract_records_from_sheet(cls, sheet: Any) -> tuple[list[dict[str, Any]], int]:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return [], -1

        best_index = -1
        best_score = -1
        best_headers: list[str] = []
        best_offset = 0
        for idx, row in enumerate(rows[:12]):
            offset = cls._leading_empty_count(list(row))
            trimmed = list(row)[offset:]
            headers = [str(cell).strip() if cell is not None else "" for cell in trimmed]
            if not any(headers):
                continue
            score = cls._header_match_score(headers)
            if score > best_score:
                best_score = score
                best_index = idx
                best_headers = headers
                best_offset = offset

        if best_index < 0 or best_score < 3:
            return [], best_score

        records: list[dict[str, Any]] = []
        for row in rows[best_index + 1 :]:
            values = list(row)[best_offset:]
            if len(values) < len(best_headers):
                values.extend([None] * (len(best_headers) - len(values)))
            if not any(value is not None and str(value).strip() for value in values):
                continue
            records.append({best_headers[idx]: values[idx] for idx in range(len(best_headers)) if best_headers[idx]})
        return records, best_score

    @staticmethod
    def _leading_empty_count(values: list[Any]) -> int:
        count = 0
        for value in values:
            if value in (None, ""):
                count += 1
                continue
            break
        return count

    @classmethod
    def _header_match_score(cls, headers: list[str]) -> int:
        normalized = {cls._normalize_key(header) for header in headers if header}
        score = 0
        for aliases in cls.COLUMN_ALIASES.values():
            if any(cls._normalize_key(alias) in normalized for alias in aliases):
                score += 1
        return score

    @classmethod
    def _normalize_row(cls, row: dict[str, Any]) -> dict[str, Any]:
        normalized_keys = {cls._normalize_key(k): v for k, v in row.items() if k}

        def get_value(field: str) -> Any:
            for alias in cls.COLUMN_ALIASES.get(field, []):
                value = normalized_keys.get(cls._normalize_key(alias))
                if value not in (None, ""):
                    return value
            return None

        level = cls._normalize_level(get_value("level"))
        report_start = cls._parse_date(get_value("date"))
        report_end = cls._parse_date(get_value("date_end"))
        day = report_start or report_end

        campaign_name = str(get_value("campaign_name") or "").strip()
        adset_name = str(get_value("adset_name") or "").strip()
        ad_name = str(get_value("ad_name") or "").strip()

        if adset_name.lower() == "all":
            adset_name = ""
        if level == "account":
            campaign_name = ""
            adset_name = ""
            ad_name = ""
        elif level in {"campaign", "adset"}:
            ad_name = ""

        campaign_id = str(get_value("campaign_id") or cls._synthetic_id("campaign", campaign_name)) if campaign_name else ""
        adset_id = str(get_value("adset_id") or cls._synthetic_id("adset", f"{campaign_id}:{adset_name}")) if adset_name else ""
        ad_id = str(get_value("ad_id") or cls._synthetic_id("ad", f"{adset_id}:{ad_name}")) if ad_name else ""

        spend = cls._to_float(get_value("spend"))
        impressions = cls._to_int(get_value("impressions"))
        clicks = cls._to_int(get_value("clicks"))
        reach = cls._to_int(get_value("reach"))
        result_type = str(get_value("result_type") or "").lower()
        raw_conversions = cls._to_int(get_value("conversions"))
        conversions = raw_conversions if cls._is_conversion_result_type(result_type) else 0
        conversion_value = cls._to_float(get_value("conversion_value"))
        ctr = cls._to_float(get_value("ctr")) or (100.0 * clicks / impressions if impressions > 0 else 0.0)
        cpc = cls._to_float(get_value("cpc")) or (spend / clicks if clicks > 0 else 0.0)
        cpm = cls._to_float(get_value("cpm")) or (1000.0 * spend / impressions if impressions > 0 else 0.0)
        frequency = cls._to_float(get_value("frequency")) or (impressions / reach if reach > 0 else 0.0)
        roas = cls._to_float(get_value("roas")) or (conversion_value / spend if spend > 0 else 0.0)

        return {
            "date": day,
            "report_start": report_start,
            "report_end": report_end,
            "campaign_id": campaign_id,
            "campaign_name": campaign_name,
            "adset_id": adset_id,
            "adset_name": adset_name,
            "ad_id": ad_id,
            "ad_name": ad_name,
            "impressions": impressions,
            "clicks": clicks,
            "spend": spend,
            "reach": reach,
            "ctr": ctr,
            "cpc": cpc,
            "cpm": cpm,
            "frequency": frequency,
            "conversions": conversions,
            "conversion_value": conversion_value,
            "roas": roas,
            "status": str(get_value("status") or "ACTIVE").upper(),
            "objective": str(get_value("objective") or cls._infer_objective(result_type)),
            "level": level,
        }

    @staticmethod
    def _normalize_key(value: str) -> str:
        normalized = str(value).lower().replace("_", " ")
        return re.sub(r"[^\w]+", " ", normalized, flags=re.UNICODE).strip()

    @staticmethod
    def _synthetic_id(prefix: str, seed: str) -> str:
        digest = hashlib.sha1(seed.encode("utf-8")).hexdigest()[:16]
        return f"{prefix}_{digest}"

    @staticmethod
    def _parse_date(value: Any) -> date | None:
        if value is None or value == "":
            return None
        if isinstance(value, date) and not isinstance(value, datetime):
            return value
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, (int, float)):
            try:
                return from_excel(value).date()
            except Exception:
                return None

        raw = str(value).strip()
        if not raw:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d", "%d.%m.%Y"):
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                continue
        try:
            return date.fromisoformat(raw)
        except ValueError:
            return None

    @staticmethod
    def _to_float(value: Any) -> float:
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)

        cleaned = str(value).strip().replace("%", "").replace("$", "").replace("\u00a0", "").replace(" ", "")
        if "," in cleaned and "." not in cleaned and cleaned.count(",") == 1:
            cleaned = cleaned.replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
        if not cleaned:
            return 0.0
        try:
            return float(cleaned)
        except ValueError:
            return 0.0

    @staticmethod
    def _to_int(value: Any) -> int:
        return int(round(CsvImportService._to_float(value)))

    @staticmethod
    def _infer_objective(result_type: str) -> str:
        value = result_type.lower()
        if any(token in value for token in ["purchase", "конверс", "lead", "install"]):
            return "CONVERSIONS"
        if any(token in value for token in ["reach", "охват"]):
            return "AWARENESS"
        if any(token in value for token in ["click", "traffic", "клик"]):
            return "TRAFFIC"
        return "UNKNOWN"

    @staticmethod
    def _is_conversion_result_type(result_type: str) -> bool:
        value = result_type.lower()
        if not value:
            return False
        if any(token in value for token in ["reach", "охват", "impression", "показ", "view", "просмотр"]):
            return False
        return any(token in value for token in ["purchase", "конверс", "lead", "install", "result", "покуп"])

    @staticmethod
    def _normalize_level(value: Any) -> str:
        normalized = CsvImportService._normalize_key(str(value or ""))
        if normalized in {"campaign", "кампания"}:
            return "campaign"
        if normalized in {"adset", "ad set", "группа объявлений"}:
            return "adset"
        if normalized in {"ad", "объявление"}:
            return "ad"
        if normalized in {"account", "аккаунт"}:
            return "account"
        return "unknown"

    @classmethod
    def _classify_report_type(cls, rows: list[dict[str, Any]]) -> str:
        if not rows:
            return "unknown"
        normalized_rows = [{cls._normalize_key(k): v for k, v in row.items() if k} for row in rows[:50]]
        has_period_bounds = any(cls._normalize_key("дата начала отчетности") in row or cls._normalize_key("reporting starts") in row for row in normalized_rows) and any(
            cls._normalize_key("дата окончания отчетности") in row or cls._normalize_key("reporting ends") in row for row in normalized_rows
        )
        explicit_date_rows = any(cls._normalize_key("date") in row or cls._normalize_key("day") in row for row in normalized_rows)
        return "period_aggregate" if has_period_bounds and not explicit_date_rows else "daily_breakdown"

    @staticmethod
    def _detect_mime_from_bytes(content_bytes: bytes) -> str:
        """Detect MIME type from magic bytes — no native library required."""
        if content_bytes[:4] == b"PK\x03\x04":
            return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        try:
            content_bytes[:2048].decode("utf-8")
            return "text/plain"
        except UnicodeDecodeError:
            try:
                content_bytes[:2048].decode("latin-1")
                return "text/plain"
            except UnicodeDecodeError:
                return "application/octet-stream"

    @staticmethod
    def _detect_extension(filename: str) -> str:
        lower = (filename or "").lower()
        if "." not in lower:
            return ""
        return lower[lower.rfind(".") :]

    @staticmethod
    def _clear_account_data(db: Session, ad_account_id: str) -> None:
        db.query(DailyAdInsight).filter(DailyAdInsight.ad_account_id == ad_account_id).delete()
        db.query(DailyAdSetInsight).filter(DailyAdSetInsight.ad_account_id == ad_account_id).delete()
        db.query(DailyCampaignInsight).filter(DailyCampaignInsight.ad_account_id == ad_account_id).delete()
        db.query(DailyAccountInsight).filter(DailyAccountInsight.ad_account_id == ad_account_id).delete()
        db.query(Ad).filter(Ad.ad_account_id == ad_account_id).delete()
        db.query(AdSet).filter(AdSet.ad_account_id == ad_account_id).delete()
        db.query(Campaign).filter(Campaign.ad_account_id == ad_account_id).delete()
        db.query(Creative).filter(Creative.ad_account_id == ad_account_id).delete()
        db.flush()

    @staticmethod
    def _upsert_campaigns(db: Session, ad_account_id: str, rows: list[dict[str, Any]]) -> dict[str, Campaign]:
        existing = {item.meta_campaign_id: item for item in db.query(Campaign).filter(Campaign.ad_account_id == ad_account_id).all()}
        result: dict[str, Campaign] = {}
        for row in rows:
            if not row["campaign_id"] or not row["campaign_name"]:
                continue
            campaign = existing.get(row["campaign_id"])
            if campaign is None:
                campaign = Campaign(
                    ad_account_id=ad_account_id,
                    meta_campaign_id=row["campaign_id"],
                    name=row["campaign_name"] or row["campaign_id"],
                    status=row["status"] or "ACTIVE",
                    objective=row["objective"] or "UNKNOWN",
                )
                db.add(campaign)
                existing[row["campaign_id"]] = campaign
            else:
                campaign.name = row["campaign_name"] or campaign.name
                campaign.status = row["status"] or campaign.status
                campaign.objective = row["objective"] or campaign.objective
            campaign.synced_at = datetime.now(timezone.utc)
            result[row["campaign_id"]] = campaign
        db.flush()
        return result

    @staticmethod
    def _upsert_adsets(db: Session, ad_account_id: str, rows: list[dict[str, Any]], campaigns: dict[str, Campaign]) -> dict[str, AdSet]:
        existing = {item.meta_adset_id: item for item in db.query(AdSet).filter(AdSet.ad_account_id == ad_account_id).all()}
        result: dict[str, AdSet] = {}
        for row in rows:
            if not row["adset_id"] or not row["adset_name"]:
                continue
            campaign = campaigns.get(row["campaign_id"])
            if campaign is None:
                continue
            ad_set = existing.get(row["adset_id"])
            if ad_set is None:
                ad_set = AdSet(
                    campaign_id=campaign.id,
                    ad_account_id=ad_account_id,
                    meta_adset_id=row["adset_id"],
                    meta_campaign_id=campaign.meta_campaign_id,
                    name=row["adset_name"] or row["adset_id"],
                    status=row["status"] or "ACTIVE",
                    optimization_goal=row["objective"] or None,
                )
                db.add(ad_set)
                existing[row["adset_id"]] = ad_set
            else:
                ad_set.campaign_id = campaign.id
                ad_set.meta_campaign_id = campaign.meta_campaign_id
                ad_set.name = row["adset_name"] or ad_set.name
                ad_set.status = row["status"] or ad_set.status
                ad_set.optimization_goal = row["objective"] or ad_set.optimization_goal
            ad_set.synced_at = datetime.now(timezone.utc)
            result[row["adset_id"]] = ad_set
        db.flush()
        return result

    @staticmethod
    def _upsert_ads(db: Session, ad_account_id: str, rows: list[dict[str, Any]], adsets: dict[str, AdSet]) -> None:
        existing = {item.meta_ad_id: item for item in db.query(Ad).filter(Ad.ad_account_id == ad_account_id).all()}
        for row in rows:
            if not row["ad_id"] or not row["adset_id"]:
                continue
            ad_set = adsets.get(row["adset_id"])
            if ad_set is None:
                continue
            ad = existing.get(row["ad_id"])
            if ad is None:
                ad = Ad(
                    ad_set_id=ad_set.id,
                    ad_account_id=ad_account_id,
                    meta_ad_id=row["ad_id"],
                    meta_adset_id=ad_set.meta_adset_id,
                    name=row["ad_name"] or row["ad_id"],
                    status="ACTIVE",
                )
                db.add(ad)
                existing[row["ad_id"]] = ad
            else:
                ad.ad_set_id = ad_set.id
                ad.meta_adset_id = ad_set.meta_adset_id
                ad.name = row["ad_name"] or ad.name
            ad.synced_at = datetime.now(timezone.utc)
        db.flush()

    @staticmethod
    def _upsert_insights(db: Session, ad_account_id: str, rows: list[dict[str, Any]], campaigns: dict[str, Campaign], adsets: dict[str, AdSet]) -> int:
        account_acc = defaultdict(lambda: _InsightAccumulator())
        campaign_acc = defaultdict(lambda: _InsightAccumulator())
        adset_acc = defaultdict(lambda: _InsightAccumulator())
        ad_acc = defaultdict(lambda: _InsightAccumulator())

        ad_map = {item.meta_ad_id: item for item in db.query(Ad).filter(Ad.ad_account_id == ad_account_id).all()}

        for row in rows:
            day = row["date"]
            if day is None:
                continue
            level = row.get("level", "unknown")
            if level == "account":
                account_acc[day].add(row)

            campaign = campaigns.get(row["campaign_id"])
            if campaign is not None and level in {"campaign", "unknown"}:
                campaign_acc[(campaign.id, day)].add(row)

            ad_set = adsets.get(row["adset_id"])
            if ad_set is not None and level in {"adset", "unknown"}:
                adset_acc[(ad_set.id, day)].add(row)

            ad = ad_map.get(row["ad_id"])
            if ad is not None and level in {"ad", "unknown"}:
                ad_acc[(ad.id, day)].add(row)

        if not account_acc:
            fallback_rows = [row for row in rows if row.get("level") in {"campaign", "unknown"} and row.get("campaign_id")]
            if not fallback_rows:
                fallback_rows = [row for row in rows if row.get("level") in {"adset", "unknown"} and row.get("adset_id")]
            for row in fallback_rows:
                if row["date"] is not None:
                    account_acc[row["date"]].add(row)

        existing_account = {item.date: item for item in db.query(DailyAccountInsight).filter(DailyAccountInsight.ad_account_id == ad_account_id).all()}
        for day, acc in account_acc.items():
            row = existing_account.get(day) or DailyAccountInsight(ad_account_id=ad_account_id, date=day)
            _fill_insight_row(row, acc.as_dict())
            db.add(row)

        existing_campaign = {
            (item.campaign_id, item.date): item
            for item in db.query(DailyCampaignInsight).filter(DailyCampaignInsight.ad_account_id == ad_account_id).all()
        }
        for key, acc in campaign_acc.items():
            campaign_id, day = key
            row = existing_campaign.get(key) or DailyCampaignInsight(ad_account_id=ad_account_id, campaign_id=campaign_id, date=day)
            _fill_insight_row(row, acc.as_dict())
            db.add(row)

        existing_adset = {
            (item.ad_set_id, item.date): item
            for item in db.query(DailyAdSetInsight).filter(DailyAdSetInsight.ad_account_id == ad_account_id).all()
        }
        for key, acc in adset_acc.items():
            ad_set_id, day = key
            row = existing_adset.get(key) or DailyAdSetInsight(ad_account_id=ad_account_id, ad_set_id=ad_set_id, date=day)
            _fill_insight_row(row, acc.as_dict())
            db.add(row)

        existing_ad = {
            (item.ad_id, item.date): item
            for item in db.query(DailyAdInsight).filter(DailyAdInsight.ad_account_id == ad_account_id).all()
        }
        for key, acc in ad_acc.items():
            ad_id, day = key
            row = existing_ad.get(key) or DailyAdInsight(ad_account_id=ad_account_id, ad_id=ad_id, date=day)
            _fill_insight_row(row, acc.as_dict())
            db.add(row)

        db.flush()
        return len(account_acc) + len(campaign_acc) + len(adset_acc) + len(ad_acc)


class _InsightAccumulator:
    def __init__(self) -> None:
        self.impressions = 0
        self.clicks = 0
        self.spend = 0.0
        self.reach = 0
        self.conversions = 0
        self.conversion_value = 0.0
        self._ctr_weighted = 0.0

    def add(self, row: dict[str, Any]) -> None:
        impressions = int(row["impressions"] or 0)
        clicks = int(row["clicks"] or 0)
        spend = float(row["spend"] or 0.0)
        reach = int(row["reach"] or 0)
        conversions = int(row["conversions"] or 0)
        conversion_value = float(row["conversion_value"] or 0.0)
        ctr = float(row["ctr"] or 0.0)

        self.impressions += impressions
        self.clicks += clicks
        self.spend += spend
        self.reach += reach
        self.conversions += conversions
        self.conversion_value += conversion_value
        self._ctr_weighted += ctr * impressions

    def as_dict(self) -> dict[str, Any]:
        ctr = self._ctr_weighted / max(self.impressions, 1)
        cpc = self.spend / self.clicks if self.clicks > 0 else 0.0
        cpm = (self.spend * 1000.0) / self.impressions if self.impressions > 0 else 0.0
        frequency = self.impressions / self.reach if self.reach > 0 else 0.0
        roas = self.conversion_value / self.spend if self.spend > 0 else 0.0
        return {
            "impressions": self.impressions,
            "clicks": self.clicks,
            "spend": round(self.spend, 4),
            "reach": self.reach,
            "ctr": round(ctr, 4),
            "cpc": round(cpc, 4),
            "cpm": round(cpm, 4),
            "frequency": round(frequency, 4),
            "conversions": self.conversions,
            "conversion_value": round(self.conversion_value, 4),
            "roas": round(roas, 4),
        }


def _fill_insight_row(row: Any, metrics: dict[str, Any]) -> None:
    row.impressions = metrics["impressions"]
    row.clicks = metrics["clicks"]
    row.spend = metrics["spend"]
    row.reach = metrics["reach"]
    row.ctr = metrics["ctr"]
    row.cpc = metrics["cpc"]
    row.cpm = metrics["cpm"]
    row.frequency = metrics["frequency"]
    row.conversions = metrics["conversions"]
    row.conversion_value = metrics["conversion_value"]
    row.roas = metrics["roas"]
    row.synced_at = datetime.now(timezone.utc)
